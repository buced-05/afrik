"""
Utilitaires d'export pour les données admin
Support CSV, Excel, PDF
"""
import csv
import json
from io import BytesIO, StringIO
from django.http import HttpResponse
from django.utils import timezone
from datetime import datetime


def export_to_csv(modeladmin, request, queryset, fields=None, filename_prefix='export'):
    """
    Exporte un queryset en CSV
    """
    if fields is None:
        fields = [field.name for field in queryset.model._meta.fields]
    
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename_prefix}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    # Ajouter BOM pour Excel UTF-8
    response.write('\ufeff')
    
    writer = csv.writer(response, delimiter=';')
    
    # En-têtes
    headers = []
    for field_name in fields:
        try:
            field = queryset.model._meta.get_field(field_name)
            headers.append(field.verbose_name or field_name)
        except:
            headers.append(field_name)
    writer.writerow(headers)
    
    # Données
    for obj in queryset:
        row = []
        for field_name in fields:
            try:
                value = getattr(obj, field_name)
                if value is None:
                    value = ''
                elif isinstance(value, (list, dict)):
                    value = json.dumps(value, ensure_ascii=False)
                elif hasattr(value, '__str__'):
                    value = str(value)
                row.append(value)
            except:
                row.append('')
        writer.writerow(row)
    
    return response


def export_to_excel(modeladmin, request, queryset, fields=None, filename_prefix='export'):
    """
    Exporte un queryset en Excel
    Nécessite openpyxl
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse(
            "Le module openpyxl n'est pas installé. Installez-le avec: pip install openpyxl",
            status=500
        )
    
    if fields is None:
        fields = [field.name for field in queryset.model._meta.fields]
    
    wb = Workbook()
    ws = wb.active
    ws.title = filename_prefix
    
    # Style pour les en-têtes
    header_fill = PatternFill(start_color="667eea", end_color="764ba2", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # En-têtes
    headers = []
    for field_name in fields:
        try:
            field = queryset.model._meta.get_field(field_name)
            headers.append(field.verbose_name or field_name)
        except:
            headers.append(field_name)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Données
    for row_num, obj in enumerate(queryset, 2):
        for col_num, field_name in enumerate(fields, 1):
            try:
                value = getattr(obj, field_name)
                if value is None:
                    value = ''
                elif isinstance(value, (list, dict)):
                    value = json.dumps(value, ensure_ascii=False)
                elif hasattr(value, '__str__'):
                    value = str(value)
                ws.cell(row=row_num, column=col_num, value=value)
            except:
                ws.cell(row=row_num, column=col_num, value='')
    
    # Ajuster la largeur des colonnes
    for col_num in range(1, len(fields) + 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 20
    
    # Sauvegarder dans un buffer
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename_prefix}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    return response


def export_to_pdf(modeladmin, request, queryset, fields=None, filename_prefix='export'):
    """
    Exporte un queryset en PDF
    Nécessite reportlab
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
    except ImportError:
        return HttpResponse(
            "Le module reportlab n'est pas installé. Installez-le avec: pip install reportlab",
            status=500
        )
    
    if fields is None:
        fields = [field.name for field in queryset.model._meta.fields]
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#667eea'),
        spaceAfter=30,
        alignment=1  # Center
    )
    
    # Titre
    title = Paragraph(f"Export {filename_prefix}", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))
    
    # En-têtes
    headers = []
    for field_name in fields:
        try:
            field = queryset.model._meta.get_field(field_name)
            headers.append(field.verbose_name or field_name)
        except:
            headers.append(field_name)
    
    # Données
    data = [headers]
    for obj in queryset:
        row = []
        for field_name in fields:
            try:
                value = getattr(obj, field_name)
                if value is None:
                    value = ''
                elif isinstance(value, (list, dict)):
                    value = json.dumps(value, ensure_ascii=False)
                elif hasattr(value, '__str__'):
                    value = str(value)
                # Limiter la longueur pour le PDF
                if len(str(value)) > 100:
                    value = str(value)[:100] + '...'
                row.append(str(value))
            except:
                row.append('')
        data.append(row)
    
    # Créer le tableau
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        # En-têtes
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        # Données
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
    ]))
    
    elements.append(table)
    
    # Construire le PDF
    doc.build(elements)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer.read(),
        content_type='application/pdf'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename_prefix}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    
    return response


def export_to_json(modeladmin, request, queryset, fields=None, filename_prefix='export'):
    """
    Exporte un queryset en JSON
    """
    if fields is None:
        fields = [field.name for field in queryset.model._meta.fields]
    
    data = []
    for obj in queryset:
        item = {}
        for field_name in fields:
            try:
                value = getattr(obj, field_name)
                if hasattr(value, 'isoformat'):  # datetime
                    value = value.isoformat()
                elif isinstance(value, (list, dict)):
                    pass  # Garder tel quel
                elif hasattr(value, '__str__'):
                    value = str(value)
                item[field_name] = value
            except:
                item[field_name] = None
        data.append(item)
    
    response = HttpResponse(
        json.dumps(data, ensure_ascii=False, indent=2, default=str),
        content_type='application/json; charset=utf-8'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename_prefix}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json"'
    
    return response

